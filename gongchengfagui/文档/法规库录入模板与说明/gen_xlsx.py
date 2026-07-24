# -*- coding: utf-8 -*-
# gen_xlsx.py — 由 2法规库 各 md 的 Front Matter 重新生成「法规清单.xlsx」
# 放在 2法规库/ 下，由 rebuild.bat 调用；自动包含所有新增法规，并保留原「备注」列。
import os, re, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))          # 2法规库
XLSX = os.path.join(BASE, '法规清单.xlsx')
SKIP = {'格式说明.md', 'README.md'}

LEVEL_ORDER = {'法律': 0, '司法解释': 1, '中央行政法规': 2, '中央部门规章': 3,
               '中央规范性文件': 4, '地方行政法规': 5, '地方规章': 6,
               '地方规范性文件': 7, '标准规范': 8, '司法案例': 9, '行政案例': 10, '政策解读': 11}

HEADERS = ['序号', '法规名称', '效力层级', '文档类型', '发布机构', '文号', '专业领域',
           '发布日期', '实施日期', '状态', '替代关系', '官方来源URL', '本地文件名', '期次', '备注']

FM_RE = re.compile(r'^---\r?\n(.*?)(\r?\n)?---', re.DOTALL)


def parse_fm(path):
    try:
        t = open(path, encoding='utf-8').read()
    except Exception:
        return None
    m = FM_RE.match(t)
    if not m:
        return None
    d = {}
    for line in m.group(1).split('\n'):
        if ':' in line:
            k, v = line.split(':', 1)
            d[k.strip()] = v.strip()
    return d


def clean(v):
    if v is None:
        return ''
    v = str(v).strip()
    if v in ("''", '""', 'None'):
        return ''
    return v


def main():
    # 收集所有法规（每个 md 一部）
    laws = []
    cat_dirs = []
    for name in sorted(os.listdir(BASE)):
        full = os.path.join(BASE, name)
        if os.path.isdir(full) and re.match(r'^\d{2}-', name):
            cat_dirs.append((name, re.sub(r'^\d{2}-', '', name), full))

    for folder, exp_level, full in cat_dirs:
        for root, _, files in os.walk(full):
            for fn in files:
                if not fn.lower().endswith('.md') or fn in SKIP:
                    continue
                fp = os.path.join(root, fn)
                rel = os.path.relpath(fp, BASE).replace('/', '\\')
                fm = parse_fm(fp)
                if not fm or not fm.get('title'):
                    continue
                laws.append({
                    '法规名称': clean(fm.get('title')),
                    '效力层级': clean(fm.get('level')) or exp_level,
                    '文档类型': clean(fm.get('doc_type')),
                    '发布机构': clean(fm.get('publisher')) or clean(fm.get('issuer')),
                    '文号': clean(fm.get('doc_number')),
                    '专业领域': clean(fm.get('field')),
                    'region': clean(fm.get('region')),
                    '发布日期': clean(fm.get('publish_date')),
                    '实施日期': clean(fm.get('effective_date')),
                    '状态': clean(fm.get('status')) or '现行',
                    '替代关系': clean(fm.get('superseded_by')),
                    '官方来源URL': clean(fm.get('source_url')),
                    '本地文件名': rel,
                    '期次': clean(fm.get('phase')) or '首期',
                    'note': clean(fm.get('note')),
                })

    # 读取现有 xlsx 的「备注」，按 本地文件名 / 法规名称 匹配以保留用户手填
    old_note_by_file = {}
    old_note_by_title = {}
    if os.path.exists(XLSX):
        try:
            wb0 = openpyxl.load_workbook(XLSX)
            if '法规清单' in wb0.sheetnames:
                ws0 = wb0['法规清单']
                hdr = [ws0.cell(1, c).value for c in range(1, ws0.max_column + 1)]
                try:
                    i_file = hdr.index('本地文件名') + 1
                    i_title = hdr.index('法规名称') + 1
                    i_note = hdr.index('备注') + 1
                except ValueError:
                    i_file = i_title = i_note = None
                for r in range(2, ws0.max_row + 1):
                    if i_file:
                        old_note_by_file[ws0.cell(r, i_file).value] = ws0.cell(r, i_note).value
                    if i_title:
                        old_note_by_title[ws0.cell(r, i_title).value] = ws0.cell(r, i_note).value
        except Exception as e:
            print('读取旧备注失败（将重新生成）：', e)

    # 排序：效力层级 → 法规名称
    laws.sort(key=lambda x: (LEVEL_ORDER.get(x['效力层级'], 99), x['法规名称']))

    # 写入「备注」：优先沿用旧值，否则用 md 的 note
    for x in laws:
        note = old_note_by_file.get(x['本地文件名'])
        if not note:
            note = old_note_by_title.get(x['法规名称'])
        if not note:
            note = x['note']
        x['备注'] = note

    # 生成工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '法规清单'

    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='305496')
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for i, x in enumerate(laws, 1):
        row = [i] + [x[h] for h in HEADERS[1:]]
        ws.append(row)
        for c in range(1, len(HEADERS) + 1):
            ws.cell(i + 1, c).border = border

    # 列宽
    widths = [6, 42, 14, 10, 22, 20, 10, 13, 13, 10, 16, 40, 34, 8, 20]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:%s%d' % (openpyxl.utils.get_column_letter(len(HEADERS)), len(laws) + 1)

    # 统计 sheet
    st = wb.create_sheet('统计')
    st['A1'] = '工程建设法规清单统计'
    st['A1'].font = Font(bold=True, size=13)
    st.append([])
    st.append(['按效力层级', '数量'])
    for lv in sorted(LEVEL_ORDER, key=lambda k: LEVEL_ORDER[k]):
        cnt = sum(1 for x in laws if x['效力层级'] == lv)
        if cnt:
            st.append([lv, cnt])
    st.append([])
    st.append(['按状态', '数量'])
    for stt in ['现行', '已废止', '待核实']:
        cnt = sum(1 for x in laws if x['状态'] == stt)
        if cnt:
            st.append([stt, cnt])
    st.append([])
    st.append(['总计', len(laws)])
    regions = set(x['region'] for x in laws if x['效力层级'].startswith('地方') and x['region'])
    st.append(['地方类覆盖地区数', len(regions)])
    st.column_dimensions['A'].width = 20
    st.column_dimensions['B'].width = 10

    wb.save(XLSX)
    print('法规清单已更新：%d 部法规，已写入 %s（含「法规清单」「统计」两个 sheet，备注列已尽量保留）。'
          % (len(laws), os.path.basename(XLSX)))
    return 0


if __name__ == '__main__':
    sys.exit(main())
